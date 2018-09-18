import openpyxl
from tablib.compat import unicode

get_column_letter = openpyxl.utils.get_column_letter

def dset_sheet(dataset, ws, freeze_panes=True):
    """Completes given worksheet from given Dataset."""
    _package = dataset._package(dicts=False)

    for i, sep in enumerate(dataset._separators):
        _offset = i
        _package.insert((sep[0] + _offset), (sep[1],))

    bold = openpyxl.styles.Font(bold=True)
    wrap_text = openpyxl.styles.Alignment(wrap_text=True)

    for i, row in enumerate(_package):
        row_number = i + 1
        for j, col in enumerate(row):
            col_idx = get_column_letter(j + 1)
            cell = ws.cell('%s%s' % (col_idx, row_number))

            # bold headers
            if (row_number == 1) and dataset.headers:
                # cell.value = unicode('%s' % col, errors='ignore')
                cell.value = unicode(col)
                cell.font = bold
                if freeze_panes:
                    #  Export Freeze only after first Line
                    ws.freeze_panes = 'A2'
                    
            # bold separators
            elif len(row) < dataset.width:
                cell.value = unicode('%s' % col, errors='ignore')
                cell.font = bold

            # wrap the rest
            else:
                try:
                    if isinstance(col, float) or isinstance(col, int):
                        cell.value = col
                    elif '\n' in col:
                        cell.value = unicode('%s' % col, errors='ignore')
                        cell.alignment = wrap_text
                    else:
                        cell.value = unicode('%s' % col, errors='ignore')
                except TypeError:
                    cell.value = unicode(col)

