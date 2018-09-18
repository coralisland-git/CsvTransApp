import os
import sys
import csv
import json
import string
import importlib
import xlrd
from xlrd.sheet import ctype_text
from openpyxl import Workbook, load_workbook
from optparse import OptionParser


class SpreadsheetReader():
    '''
    Unified Interface to read rows from types - .xls, .xlsx and .csv
    '''

    def __init__(self, filename):
        self.filename = filename
        self.filename_prefix, self.type = os.path.splitext(filename)

    def get_rows(self):
        rows = []

        if self.type == '.csv':
            with open(self.filename, newline='') as csvfile:
                reader = csv.reader(csvfile, delimiter=',', quotechar='|')
                for row in reader:
                    rows.append(row)

        if self.type == '.xlsx':
            wb = load_workbook(self.filename)
            ws = wb.worksheets[0]
            for row in ws.iter_rows():
                rows.append([cell.value for cell in row])

        if self.type == '.xls':
            book = xlrd.open_workbook(self.filename)
            sheet = book.sheet_by_index(0)

            rows = []

            for i in range(sheet.nrows):
                row = []
                for cell in sheet.row(i):
                    if ctype_text.get(cell.ctype, 'unknown type') == 'xldate':
                        value = xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
                    else:
                        value = cell.value

                    row.append(value)
                rows.append(row)

        return rows



class SpreadsheetWriter():
    '''
    Create .xlsx files from rows data
    rows data is a 2D array - like the data that SpreadsheetReader.get_rows() returns
    '''

    def __init__(self, filename, output_sheet_name="Sheet1"):
        self.filename_prefix, self.type = os.path.splitext(filename)
        self.wb = Workbook()
        self.output_sheet_name = output_sheet_name

    def set_rows(self, rows):
        ws = self.wb.worksheets[0]
        ws.title = self.output_sheet_name
        for row in rows:
            ws.append(row)

    def save_file(self):
        self.wb.save(filename = self.filename_prefix + '.xlsx')



class FormatterOptions:
    '''
    Use to hold the 'options' to Formatter() constructor with support
    to access properties like 'options.input_path'
    '''

    format = None
    input_path = None
    output_sheet_name = None
    def __init__(self, format, input_path, output_sheet_name="Sheet1"):
        self.format = format
        self.input_path = input_path
        self.output_sheet_name = output_sheet_name



class Formatter():
    '''
    Format .csv/.xls/.xlsx files into desired format and save as .xlsx

    options.input_path: the file that has to be formatted

    options.format: the format of output file

    - A JSON file is expected corresponding to the specified format in
    <root_dir>/formats directory. If you specify the format as xyz,
    the program would expect a format specification JSON file -
    <root_dir>/formats/xyz.json
    
    options.output_sheet_name: the name of the sheet in the output file
    '''

    def __init__(self, options):
        self.options = options
        
        # The input file data        
        self.in_spreadsheet = SpreadsheetReader(self.options.input_path)

        # The output file name
        self.filename_prefix, self.type = os.path.splitext(self.options.input_path)
        out_filename = self.filename_prefix + '_formatted' + self.type

        if not self.options.output_sheet_name:
            self.options.output_sheet_name = "Sheet1"

        # The output file data
        self.out_spreadsheet = SpreadsheetWriter(out_filename, output_sheet_name=self.options.output_sheet_name)

        # The format configuration
        self.c = {}

        # 'header_row' and 'data' together is the whole file data
        self.header_row = []
        self.data = []

        # operations_module is a python file in the <root_dir>/operations directory
        # which contains the custom operations specific to a format
        try:
            self.operations_module = importlib.import_module('operations.' + self.options.format)
        except Exception as e:
            try:
                # Fixme: This inner try-except is there to support testing
                # Could find a way to avoid this
                parentPath = os.path.abspath("..")
                if parentPath not in sys.path:
                    sys.path.insert(0, parentPath)
                self.operations_module = importlib.import_module('operations.' + self.options.format.split(os.sep)[-1])
            except Exception as e:
                pass    
 

    def _drop_columns(self):
        '''
        Delete columns
        '''
        if 'columns' not in self.c:
            return

        drop_cols = [col2num(col) for col in self.c['columns'] if self.c['columns'][col]=='drop']

        #Required because: After deletion of col #n, we have col #n+1 now at #n position
        drop_cols.sort(reverse=True)

        for row in self.data:
            for col in drop_cols:
                del row[col]

        for col in drop_cols:
            del self.header_row[col]


    def _create_new_columns(self):
        '''
        Add new columns at the end
        '''

        if 'new_columns' not in self.c:
            return

        replace_based_on_cols = {}
        concat_based_on_cols = {}
        new_header_titles = []

        for new_col_def in self.c['new_columns']:
            if 'action' in self.c['new_columns'][new_col_def] and self.c['new_columns'][new_col_def]['action'] == "replace":
                    if 'based_on' in self.c['new_columns'][new_col_def]:
                        replace_based_on_cols[new_col_def] = self.c['new_columns'][new_col_def]
                        if 'has_header_row' in self.c and self.c['has_header_row'] and self.header_row:
                            new_header_titles.append(self.c['new_columns'][new_col_def]['header_title'])
            if 'action' in self.c['new_columns'][new_col_def] and self.c['new_columns'][new_col_def]['action'] == "concatenate":
                    if 'based_on' in self.c['new_columns'][new_col_def]:
                        concat_based_on_cols[new_col_def] = self.c['new_columns'][new_col_def]
                        if 'has_header_row' in self.c and self.c['has_header_row'] and self.header_row:
                            new_header_titles.append(self.c['new_columns'][new_col_def]['header_title'])

        for row in self.data:
            for col in replace_based_on_cols:
                try:
                    row.append(replace_based_on_cols[col]['with'][str(row[col2num(replace_based_on_cols[col]['based_on'])])])
                except Exception as e:
                    # sometimes while reading from csv integers are read as floats (20 as 20.0), this could cause dict KeyErrors
                    if isfloat(str(row[col2num(replace_based_on_cols[col]['based_on'])])):
                        row.append(replace_based_on_cols[col]['with'][str(int(row[col2num(replace_based_on_cols[col]['based_on'])]))])
                    else:
                        print(e)
                        sys.exit(-1)
            for col in concat_based_on_cols:
                row.append(concat_based_on_cols[col]['join_string'].join([str(row[col2num(c)]) for c in concat_based_on_cols[col]['based_on']]))

                    
        for header_title in new_header_titles:
            self.header_row.append(header_title)


    def _process_columns(self):
        '''
        Perform operations based on columns
        '''
        if 'columns' not in self.c:
            return

        # Gather column-numbers for the deletion of rows based on column-uniqueness
        unique_cols = [col2num(col) for col in self.c['columns'] if self.c['columns'][col]=='unique']

        colnum_vals_map = {} #colnum: values found in the column
        rows_to_be_deleted = []

        # Gather row-numbers that have duplicates as per 'unique_cols'
        for row_num, row in enumerate(self.data):
            for col in unique_cols:
                if col in colnum_vals_map:
                    if row[col] in colnum_vals_map[col]:
                        rows_to_be_deleted.append(row_num)
                    else:
                        colnum_vals_map[col].append(row[col])
                else:
                    colnum_vals_map[col] = [row[col]]

        # Required because: After deletion of row #n, we have row #n+1 now at #n position
        rows_to_be_deleted.sort(reverse=True)

        # Delete duplicate rows
        for row_num in rows_to_be_deleted:
            del self.data[row_num]


        # Gather info for custom operations, replacements, cut-pastes and clearing
        operation_cols = {}
        name_func_map = {}
        replace_cols = {}
        replace_based_on_cols = {}
        cut_paste_cols = {}

        for col in self.c['columns']:
            if isinstance(self.c['columns'][col], dict):
                if self.c['columns'][col]['action'] == 'operation':
                    operation_cols[col2num(col)] = {'function': self.c['columns'][col]['function']}
                    try:
                        name_func_map[self.c['columns'][col]['function']] = getattr(self.operations_module, self.c['columns'][col]['function'])
                    except Exception as e:
                        print('Error while reading definition of the custom operation {}'.format(self.c['columns'][col]['function']))
                        sys.exit(-1)
                if self.c['columns'][col]['action'] == 'replace':
                    if 'based_on' in self.c['columns'][col]:
                        replace_based_on_cols[col2num(col)] = self.c['columns'][col]
                    else:
                        replace_cols[col2num(col)] = {'with': self.c['columns'][col]['with']}
                if self.c['columns'][col]['action'] == 'cutpaste':
                    cut_paste_cols[col2num(col)] = col2num(self.c['columns'][col]['from'])

        clear_cols = [col2num(col) for col in self.c['columns'] if self.c['columns'][col]=='clear']


        #Perform custom operations, replacements, cut-pastes and clearing
        for row_num, row in enumerate(self.data):

            for col in cut_paste_cols:
                row[col] = row[cut_paste_cols[col]]
                row[cut_paste_cols[col]] = ''

            for col in operation_cols:
                row[col] = name_func_map[operation_cols[col]['function']](row[col], row, row_num)

            for col in replace_cols:
                row[col] = replace_cols[col]['with'][row[col]]

            for col in replace_based_on_cols:
                try:
                    row[col] = replace_based_on_cols[col]['with'][str(row[col2num(replace_based_on_cols[col]['based_on'])])]
                except Exception as e:
                    # sometimes while reading from csv integers are read as floats (20 as 20.0), this could cause dict KeyErrors
                    if isfloat(str(row[col2num(replace_based_on_cols[col]['based_on'])])):
                        row[col] = replace_based_on_cols[col]['with'][str(int(row[col2num(replace_based_on_cols[col]['based_on'])]))]
                    else:
                        print(e)
                        sys.exit(-1)

            for col in clear_cols:
                row[col] = ''


    def _process_header(self):
        '''
        Perform operations on header
        Example: replace header-text of 'A' with 'EmpID'
        '''
        if 'has_header_row' not in self.c or not self.c['has_header_row'] or 'header_rows' not in self.c:
            return

        for col in list(self.c['header_rows'].keys()):
            instruction = self.c['header_rows'][col]
            if isinstance(instruction, dict):
                if instruction['action'] == 'replace':
                    try:
                        self.header_row[col2num(col)] = instruction['with']
                    except:
                        pass #Wasn't able to set the cell value, may be the cell is non-existent


    def _separate_header_and_body(self):
        if 'has_header_row' in self.c and self.c['has_header_row']:
            self.header_row = self.data[0]
            self.data = self.data[1:]


    def _merge_header_and_body(self):
        if 'has_header_row' in self.c and self.c['has_header_row'] and self.header_row:
            self.data.insert(0, self.header_row)


    def _process_rows(self):
        '''
        Do operations that are row-wise in nature
        Example: drop row 1, drop row 3
        Note: this is doesn't consider if there is a header row or not
        '''
        if not 'rows' in self.c:
            return

        drop_rows = [int(row) for row in self.c['rows'] if self.c['rows'][row]=='drop']

        #Required because: After deletion of row #n, we have row #n+1 now at #n position
        drop_rows.sort(reverse=True)
        
        for row in drop_rows:
            try:
                del self.data[row - 1]
            except:
                pass #Wasn't able to delete the row, may be the row is non-existent


    def _load_format(self):
        '''
        Read format specification into self.c
        '''
        if os.sep in self.options.format:
            format_fpath = self.options.format
        else:
            format_fpath = os.path.join('formats', self.options.format)

        if not format_fpath.endswith('.json'):
            format_fpath += '.json'

        with open(format_fpath) as fp:
            self.c = json.load(fp)


    def run(self):
        """
        Execute the Formatter function in steps as broken down into functions
        """

        self.data = self.in_spreadsheet.get_rows()

        self._load_format()
        self._process_rows()
        self._separate_header_and_body()
        self._process_header()
        self._process_columns()
        self._create_new_columns()
        self._drop_columns()
        self._merge_header_and_body()

        self.out_spreadsheet.set_rows(self.data)
        self.out_spreadsheet.save_file()



def col2num(col):
    '''
    Return a number corresponding to the column label
    col2num('A') will return 0
    col2num('AB') will return 27
    '''
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num - 1



def isfloat(value):
    try:
        float(value)
        return True
    except ValueError:
        return False



def process_options():
    parser = OptionParser()
    parser.add_option("-f", "--format", dest="format",help="enter format", metavar="FORMAT")
    parser.add_option("-i", "--input", dest="input_path", help="input file path", metavar="PATH")
    parser.add_option("-s", "--sheet", dest="output_sheet_name", help="sheet name", metavar="SHEET_NAME")

    (options, args) = parser.parse_args()

    if not options.format:
        parser.error('format not provided (-f option)')

    if not options.input_path:
            parser.error('input path not provided (-i option)')

    options.input_path = options.input_path.rstrip('/').rstrip('\\')

    return options



def main():
    options = process_options()
    Formatter(options).run()



if __name__== '__main__':
    main()
